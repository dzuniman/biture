from openpyxl import load_workbook
import itertools
wb = load_workbook('Quote2Cash.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    print('---', name)
    for row in itertools.islice(wb[name].iter_rows(min_row=1, max_row=10, values_only=True), 10):
        print(row)
